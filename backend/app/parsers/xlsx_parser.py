import logging
from openpyxl import load_workbook

from app.parsers.base import DocumentParser, ParsedDocument, ParsedPage

logger = logging.getLogger(__name__)


class XlsxParser(DocumentParser):
    """
    Parses .xlsx files using openpyxl.
    Each worksheet is treated as one logical 'page'. Cell values from each row
    are joined with a tab separator; rows are joined with newlines.
    """

    def parse(self, file_path: str) -> ParsedDocument:
        pages: ParsedDocument = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                rows_text: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    # Filter out entirely empty rows
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    if any(c.strip() for c in cells):
                        rows_text.append("\t".join(cells))
                sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text)
                pages.append(ParsedPage(page_number=sheet_index, raw_text=sheet_text))
            wb.close()
        except Exception as exc:
            logger.error("XlsxParser failed for %s: %s", file_path, exc)
            raise
        return pages
